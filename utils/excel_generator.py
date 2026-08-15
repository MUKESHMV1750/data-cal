"""
Excel Generator using openpyxl for student data export.
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from django.utils import timezone


PRIMARY_COLOR = '1E3A5F'
ACCENT_COLOR = '2563EB'
HEADER_BG = '1E3A5F'
HEADER_FG = 'FFFFFF'
ALT_ROW = 'EFF6FF'
BORDER_COLOR = 'D1D5DB'


def make_border(color=BORDER_COLOR):
    side = Side(style='thin', color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def generate_students_excel(students):
    """Generate a comprehensive Excel file for all/filtered students."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Student Data'

    # === Column Headers ===
    headers = [
        'S.No.', 'Student Name', 'Enrollment Number', 'Exam Reg. Number',
        'UMIS Number', 'Program of Study', 'Branch/Discipline', 'Year of Completion',
        'Email ID', 'Mobile Number', 'Alternate Mobile', 'Date of Birth',
        'Gender', 'Category', 'Blood Group',
        'PAN Number', 'Aadhaar Number', 'Passport Number',
        'Permanent Address', "Guardian's Name", "Guardian's Mobile",
        "Guardian's Occupation", 'Annual Income (₹)',
        '10th Percentage', '10th Passing Year',
        '12th Percentage', '12th Passing Year',
        'Diploma Percentage', 'Diploma Passing Year',
        'OGPA (Till 5th Sem)',
        'History of Arrears', 'Current Backlog', 'No. of Backlogs',
        'Remarks/Gap Reason',
        'Programming Languages', 'Communication Languages', 'Additional Skills',
        'Internship Company', 'Internship Dates',
        'Laptop/Desktop (Webcam&Mic)', 'High-Speed Internet',
        'Career Preference', 'Willingness to Relocate',
        'NPTEL/SWAYAM Courses', 'Future Internship (Dec26-Apr27)',
        'TPC Acceptance', 'Submitted Date', 'Submitted Time'
    ]

    # === Title Row ===
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    title_cell.value = 'ANNAMALAI UNIVERSITY — Training & Placement Cell | Student Data Report'
    title_cell.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor=PRIMARY_COLOR)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # === Generated Date Row ===
    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    now = timezone.localtime(timezone.now())
    date_cell = ws['A2']
    date_cell.value = f"Generated on: {now.strftime('%d-%m-%Y %H:%M:%S')} | Total Records: {students.count()}"
    date_cell.font = Font(name='Calibri', italic=True, size=10, color='374151')
    date_cell.fill = PatternFill('solid', fgColor='DBEAFE')
    date_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # === Header Row ===
    header_row = 3
    header_font = Font(name='Calibri', bold=True, size=10, color=HEADER_FG)
    header_fill = PatternFill('solid', fgColor=ACCENT_COLOR)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = make_border('FFFFFF')
    ws.row_dimensions[header_row].height = 40

    # === Data Rows ===
    student_list = list(students)
    bool_val = lambda v: 'Yes' if v else 'No'
    data_font = Font(name='Calibri', size=9)
    data_align_center = Alignment(horizontal='center', vertical='center')
    data_align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for idx, s in enumerate(student_list):
        row_num = header_row + 1 + idx
        alt = idx % 2 == 1
        row_fill = PatternFill('solid', fgColor=ALT_ROW) if alt else None

        aadhaar_masked = f"XXXX-XXXX-{s.aadhaar_number[-4:]}" if s.aadhaar_number and len(s.aadhaar_number) >= 4 else ''

        row_data = [
            idx + 1, s.name, s.enrollment_number, s.exam_registration_number,
            s.umis_number, s.program_of_study, s.branch, s.year_of_completion,
            s.email, s.mobile, s.alternate_mobile or '',
            s.date_of_birth.strftime('%d-%m-%Y') if s.date_of_birth else '',
            s.gender, s.category, s.blood_group,
            s.pan_number or '', aadhaar_masked, s.passport_number or '',
            s.permanent_address, s.guardian_name, s.guardian_mobile,
            s.guardian_occupation, float(s.guardian_annual_income),
            float(s.tenth_percentage), s.tenth_passing_year,
            float(s.twelfth_percentage) if s.twelfth_percentage else '',
            s.twelfth_passing_year or '',
            float(s.diploma_percentage) if s.diploma_percentage else '',
            s.diploma_passing_year or '',
            float(s.ogpa),
            bool_val(s.history_of_arrears), bool_val(s.current_backlog),
            s.number_of_backlogs, s.remarks or '',
            s.programming_languages or '', s.communication_languages or '',
            s.additional_skills or '', s.internship_company or '',
            s.internship_dates or '',
            bool_val(s.laptop_desktop), bool_val(s.high_speed_internet),
            s.career_preference, s.willingness_to_relocate,
            s.nptel_courses or '', bool_val(s.future_internship),
            'Accepted' if s.tpc_acceptance else 'Not Accepted',
            s.created_at.strftime('%d-%m-%Y') if s.created_at else '',
            s.created_at.strftime('%H:%M:%S') if s.created_at else '',
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = data_font
            cell.border = make_border()
            if col_idx in [1, 8, 12, 13, 14, 15, 24, 25, 26, 27, 28, 29, 30, 32, 33, 44, 47, 48]:
                cell.alignment = data_align_center
            else:
                cell.alignment = data_align_left
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_num].height = 22

    # === Column Widths ===
    col_widths = [
        6, 25, 20, 20, 18, 18, 30, 10, 28, 14, 14, 13,
        10, 12, 10, 14, 16, 16, 35, 25, 14, 22, 14,
        12, 12, 12, 12, 14, 12, 8,
        12, 12, 10, 30,
        30, 30, 30, 25, 20,
        18, 16, 25, 18,
        30, 14, 12, 12, 12
    ]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze panes
    ws.freeze_panes = f'A{header_row + 1}'

    # Add a summary sheet
    ws_summary = wb.create_sheet('Summary')
    ws_summary['A1'] = 'STUDENT DATA SUMMARY REPORT'
    ws_summary['A1'].font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    ws_summary['A1'].fill = PatternFill('solid', fgColor=PRIMARY_COLOR)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A1:D1')
    ws_summary.row_dimensions[1].height = 28

    from django.db.models import Count
    from students.models import Student as StudentModel

    summary_data = [
        ['Total Students', StudentModel.objects.count()],
        ['Students with Backlogs', StudentModel.objects.filter(current_backlog=True).count()],
        ['Students without Backlogs', StudentModel.objects.filter(current_backlog=False).count()],
        ['Students with Internship', StudentModel.objects.exclude(internship_company='').count()],
        ['Students with Laptop/PC', StudentModel.objects.filter(laptop_desktop=True).count()],
        ['TPC Rules Accepted', StudentModel.objects.filter(tpc_acceptance=True).count()],
    ]
    for r, (label, val) in enumerate(summary_data, 3):
        ws_summary[f'A{r}'] = label
        ws_summary[f'A{r}'].font = Font(name='Calibri', bold=True, size=10)
        ws_summary[f'B{r}'] = val
        ws_summary[f'B{r}'].font = Font(name='Calibri', size=10)
        ws_summary[f'B{r}'].alignment = Alignment(horizontal='center')
        for col in ['A', 'B']:
            ws_summary[f'{col}{r}'].border = make_border()
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
